"""Parse stocks and flows from CSV or Excel (.xlsx) files.

Expected column layout (case-insensitive headers):
  type | name | equation | initial_value | units | description

``type`` must be one of: stock, flow, aux.
``initial_value`` is required for stocks; ignored for flows/aux.
``equation`` is required for flows/aux; optional for stocks (defaults to "0").
``units`` and ``description`` are optional.

Rows whose ``type`` cell is empty or starts with ``#`` are silently skipped
(handy for section headers / comments).
"""

from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any

from app.schemas.model import (
    ModelDocument,
    ModelMetadata,
    Position,
)


_VALID_TYPES = {"stock", "flow", "aux"}

# Column aliases (all lowered) → canonical key
_COLUMN_ALIASES: dict[str, str] = {
    "type": "type",
    "kind": "type",
    "name": "name",
    "variable": "name",
    "label": "name",
    "equation": "equation",
    "formula": "equation",
    "initial_value": "initial_value",
    "initial value": "initial_value",
    "init": "initial_value",
    "units": "units",
    "unit": "units",
    "description": "description",
    "desc": "description",
    "note": "description",
}


def _normalise_name(raw: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", raw.strip().lower()).strip("_")


def _layout_positions(n: int) -> list[Position]:
    """Simple grid layout: stocks on top row, flows/aux below."""
    positions: list[Position] = []
    for i in range(n):
        col = i % 5
        row = i // 5
        positions.append(Position(x=100 + col * 200, y=100 + row * 150))
    return positions


def _parse_rows(rows: list[dict[str, str]], filename: str) -> dict[str, Any]:
    """Shared logic for CSV and Excel: rows is a list of {header: value}."""
    # Normalise headers to canonical keys
    header_map: dict[str, str] = {}
    for raw_header in (rows[0].keys() if rows else []):
        lowered = raw_header.strip().lower()
        canonical = _COLUMN_ALIASES.get(lowered)
        if canonical:
            header_map[raw_header] = canonical

    if "type" not in header_map.values():
        raise ValueError("Missing required column: 'type' (or 'kind')")
    if "name" not in header_map.values():
        raise ValueError("Missing required column: 'name' (or 'variable' / 'label')")

    nodes: list[dict[str, Any]] = []
    warnings: list[dict[str, str]] = []
    seen_names: set[str] = set()

    for row_idx, raw_row in enumerate(rows, start=2):  # 1-based, header is row 1
        mapped: dict[str, str] = {}
        for raw_key, value in raw_row.items():
            canon = header_map.get(raw_key)
            if canon:
                mapped[canon] = (value or "").strip()

        row_type = mapped.get("type", "").lower().strip()
        if not row_type or row_type.startswith("#"):
            continue

        if row_type not in _VALID_TYPES:
            warnings.append({
                "code": "SS_UNKNOWN_TYPE",
                "message": f"Row {row_idx}: unknown type '{row_type}', skipped",
                "severity": "warning",
            })
            continue

        raw_name = mapped.get("name", "").strip()
        if not raw_name:
            warnings.append({
                "code": "SS_MISSING_NAME",
                "message": f"Row {row_idx}: missing name, skipped",
                "severity": "warning",
            })
            continue

        safe_name = _normalise_name(raw_name)
        if safe_name in seen_names:
            suffix = 2
            while f"{safe_name}_{suffix}" in seen_names:
                suffix += 1
            safe_name = f"{safe_name}_{suffix}"
        seen_names.add(safe_name)

        equation = mapped.get("equation", "") or "0"
        initial_value = mapped.get("initial_value", "") or "0"
        units = mapped.get("units") or None
        description = mapped.get("description") or None

        node: dict[str, Any] = {
            "id": f"ss_{uuid.uuid4().hex[:8]}",
            "type": row_type,
            "name": safe_name,
            "label": raw_name,
            "equation": equation,
            "units": units,
            "position": {"x": 0, "y": 0},  # filled after
        }

        if description:
            node["annotation"] = {"note": description}

        if row_type == "stock":
            node["initial_value"] = initial_value
        elif row_type == "flow":
            node["source_stock_id"] = None
            node["target_stock_id"] = None

        nodes.append(node)

    # Assign layout positions
    positions = _layout_positions(len(nodes))
    for i, node in enumerate(nodes):
        node["position"] = positions[i].model_dump()

    # Build outputs: stock and flow names
    outputs = [n["name"] for n in nodes if n["type"] in ("stock", "flow")][:20]

    model_id = f"ss_{_normalise_name(filename.rsplit('.', 1)[0]) or 'model'}"
    doc = ModelDocument(
        id=model_id,
        name=filename,
        version=1,
        metadata=ModelMetadata(
            imported=None,
        ),
        nodes=nodes,
        edges=[],
        outputs=outputs,
    )

    return {"model": doc, "warnings": warnings, "node_count": len(nodes)}


def parse_csv(payload: str, filename: str) -> dict[str, Any]:
    """Parse a CSV string into a ModelDocument."""
    reader = csv.DictReader(io.StringIO(payload))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty or has no data rows")
    return _parse_rows(rows, filename)


def parse_excel(payload: bytes, filename: str) -> dict[str, Any]:
    """Parse an Excel (.xlsx) file into a ModelDocument."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl is required for Excel import; install via: pip install openpyxl")

    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel workbook has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Excel file is empty or has no header row")

    headers = [str(cell or "").strip() for cell in header_row]
    rows: list[dict[str, str]] = []
    for data_row in rows_iter:
        row_dict: dict[str, str] = {}
        for i, cell in enumerate(data_row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell) if cell is not None else ""
        rows.append(row_dict)

    wb.close()

    if not rows:
        raise ValueError("Excel file has no data rows after the header")
    return _parse_rows(rows, filename)
